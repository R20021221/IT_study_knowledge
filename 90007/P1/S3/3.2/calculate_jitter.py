"""
Section 3.2 — Mean RTT and Jitter (sample standard deviation) calculation.

For each of the 10 iperf hosts we collected 3 ping runs of 4 ICMP echoes
(N = 12 packets per host).  This script reads those 12 raw RTT values
from Section3.2_PingData.xlsx → Packets sheet, then computes:

    Mean RTT:  x̄ = (Σ xᵢ) / N
    Jitter σ:  σ = √[ Σ(xᵢ - x̄)² / (N - 1) ]      (sample standard deviation)

Distance values are copied from Section2.2_Analysis.xlsx (column K, Haversine
result).  They are written into the HostSummary sheet as plain numeric values
to make the linkage with Section 2 visible inside this workbook.

Outputs (overwritten on each run):
    - Section3.2_PingData.xlsx → adds/replaces "HostSummary" sheet
    - Section3.2_AvgRTT_vs_Distance.png
    - Section3.2_Jitter_vs_Distance.png

Usage:
    python3 calculate_jitter.py
"""

import math
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import matplotlib.pyplot as plt


HERE   = Path(__file__).resolve().parent
XLSX   = HERE / "Section3.2_PingData.xlsx"
PNG_RTT = HERE / "Section3.2_AvgRTT_vs_Distance.png"
PNG_JIT = HERE / "Section3.2_Jitter_vs_Distance.png"

# Distances from Section 2.2 → Distance vs Hops sheet, column K (Haversine).
# Listed in host order 1–10.  Reproducing them here as values keeps Section 3.2
# self-contained while making the link to Section 2 visible.
DISTANCES_KM = {
    1: 14155, 2: 16321, 3: 16321, 4: 15636, 5: 15207,
    6: 16292, 7: 17761, 8: 13621, 9: 14482, 10: 12679,
}


# ── 1. Read raw 12-packet samples per host from Packets sheet ──────────────
def load_packets():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Packets"]
    hosts = defaultdict(lambda: {"name": "", "ip": "", "rtts": []})
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Host#, Hostname, Resolved, IP, Run#, icmp_seq, bytes, ttl, time(ms)
        host_num, hostname, _resolved, ip, _run, _seq, _bytes, _ttl, time_ms = row
        if host_num is None or time_ms is None:
            continue
        hosts[host_num]["name"] = hostname
        hosts[host_num]["ip"]   = ip
        hosts[host_num]["rtts"].append(float(time_ms))
    return hosts


# ── 2. Compute mean and jitter, with a verbose breakdown ───────────────────
def compute(hosts):
    results = []
    for num in sorted(hosts.keys()):
        h    = hosts[num]
        rtts = h["rtts"]
        N    = len(rtts)

        mean_rtt    = sum(rtts) / N
        sq_diffs    = [(x - mean_rtt) ** 2 for x in rtts]
        sum_sq      = sum(sq_diffs)
        variance    = sum_sq / (N - 1)
        jitter_sd   = math.sqrt(variance)

        # Verbose step-by-step printout
        print(f"\n=== Host {num}: {h['name']} ({h['ip']}) ===")
        print(f"  Distance from Melbourne: {DISTANCES_KM[num]:,} km")
        print(f"  N = {N}")
        print(f"  RTT samples (ms): {[round(x, 3) for x in rtts]}")
        print(f"  Σxᵢ        = {sum(rtts):.3f}")
        print(f"  x̄ = Σxᵢ/N  = {sum(rtts):.3f} / {N} = {mean_rtt:.3f} ms")
        print(f"  Per-packet (xᵢ - x̄):")
        for i, x in enumerate(rtts):
            print(f"    x{i+1:>2} - x̄ = {x:>8.3f} - {mean_rtt:.3f} = {x - mean_rtt:>+8.3f}"
                  f"   →  squared = {(x - mean_rtt) ** 2:>10.4f}")
        print(f"  Σ(xᵢ - x̄)² = {sum_sq:.4f}")
        print(f"  variance   = Σ(xᵢ - x̄)² / (N - 1) = {sum_sq:.4f} / {N - 1} = {variance:.4f}")
        print(f"  σ          = √variance = √{variance:.4f} = {jitter_sd:.3f} ms")

        results.append({
            "num":      num,
            "name":     h["name"],
            "ip":       h["ip"],
            "distance": DISTANCES_KM[num],
            "mean_rtt": mean_rtt,
            "jitter":   jitter_sd,
            "N":        N,
        })
    return results


# ── 3. Write HostSummary sheet (plain numeric values) ──────────────────────
def write_host_summary(results):
    wb = load_workbook(XLSX)
    if "HostSummary" in wb.sheetnames:
        del wb["HostSummary"]
    ws = wb.create_sheet("HostSummary", 2)

    # Title / subtitle
    ws.cell(row=1, column=1,
            value="Section 3.2 — Per-Host Summary  (N = 12 packets per host = 3 runs × 4 echoes)")
    ws.cell(row=2, column=1,
            value="Mean RTT  x̄ = Σxᵢ / N    Jitter σ = √[Σ(xᵢ − x̄)² / (N − 1)]"
                  "    Distance: from Section2.2_Analysis.xlsx → 'Distance vs Hops' sheet, column K")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["Host #", "Hostname", "IP",
               "Distance (km)", "Mean RTT (ms)", "Jitter σ (ms)", "N"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)

    for i, r in enumerate(results, start=4):
        ws.cell(row=i, column=1, value=r["num"])
        ws.cell(row=i, column=2, value=r["name"])
        ws.cell(row=i, column=3, value=r["ip"])
        ws.cell(row=i, column=4, value=r["distance"])
        ws.cell(row=i, column=5, value=round(r["mean_rtt"], 3))
        ws.cell(row=i, column=6, value=round(r["jitter"], 3))
        ws.cell(row=i, column=7, value=r["N"])

    # ── Styling ──
    arial         = Font(name="Arial", size=11)
    title_font    = Font(name="Arial", size=12, bold=True, color="1F3864")
    subtitle_font = Font(name="Arial", size=10, italic=True, color="404040")
    header_font   = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill   = PatternFill("solid", start_color="305496")
    ext_link_font = Font(name="Arial", size=11, color="FF0000")  # red = from Section 2
    thin   = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = left
    ws.cell(row=2, column=1).font = subtitle_font
    ws.cell(row=2, column=1).alignment = left

    for c in range(1, 8):
        cell = ws.cell(row=3, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for r in range(4, 14):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c == 4:
                cell.font = ext_link_font          # Distance (from S2) — red
                cell.alignment = right
                cell.number_format = "#,##0"
            elif c in (5, 6):
                cell.font = arial
                cell.alignment = right
                cell.number_format = "0.000"
            elif c == 7:
                cell.font = arial
                cell.alignment = center
            else:
                cell.font = arial
                cell.alignment = center if c in (1,) else left

    # Column widths
    widths = {"A": 7, "B": 30, "C": 18, "D": 14, "E": 14, "F": 14, "G": 6}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    wb.save(XLSX)
    print(f"\n✓ HostSummary sheet written to {XLSX.name}")


# ── 4. Generate PNG charts with linear trendline + equation ────────────────
def linreg(xs, ys):
    """Ordinary least-squares slope and intercept."""
    n  = len(xs)
    mx = sum(xs) / n
    my = sum(ys) / n
    num = sum((xs[i] - mx) * (ys[i] - my) for i in range(n))
    den = sum((x - mx) ** 2 for x in xs)
    slope     = num / den
    intercept = my - slope * mx
    return slope, intercept


def make_scatter(results, y_key, y_label, title, fname):
    x = [r["distance"] for r in results]
    y = [r[y_key]      for r in results]
    names = [r["name"] for r in results]

    fig, ax = plt.subplots(figsize=(9, 5.5), dpi=150)
    ax.scatter(x, y, s=60, color="#2166ac", edgecolor="black", zorder=3)
    for xi, yi, nm in zip(x, y, names):
        ax.annotate(nm, (xi, yi), xytext=(5, 5),
                    textcoords="offset points", fontsize=8, color="#333333")

    # Linear fit + equation
    m, b = linreg(x, y)
    xs_line = [min(x) - 500, max(x) + 500]
    ys_line = [m * xi + b for xi in xs_line]
    ax.plot(xs_line, ys_line, linestyle="--", color="#888888",
            linewidth=1, label="linear fit")

    eqn = f"y = {m:+.4f} · x {b:+.2f}"
    ax.text(0.02, 0.98, f"{eqn}\nN = {len(x)}",
            transform=ax.transAxes, va="top", fontsize=10,
            bbox=dict(boxstyle="round,pad=0.4", facecolor="white", edgecolor="#888"))

    ax.set_xlabel("Distance from Melbourne (km)")
    ax.set_ylabel(y_label)
    ax.set_title(title)
    ax.grid(True, linestyle=":", linewidth=0.5, alpha=0.6)
    ax.legend(loc="lower right", fontsize=9)
    plt.tight_layout()
    plt.savefig(fname)
    plt.close()
    print(f"✓ {Path(fname).name}   (slope = {m:+.4f}, intercept = {b:+.2f})")


# ── Main ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    hosts   = load_packets()
    results = compute(hosts)

    print("\n" + "=" * 70)
    print(f"{'#':<3} {'Host':<32} {'Dist':>7} {'Mean (ms)':>11} {'σ (ms)':>9}")
    print("-" * 70)
    for r in results:
        print(f"{r['num']:<3} {r['name']:<32} {r['distance']:>7,} "
              f"{r['mean_rtt']:>11.3f} {r['jitter']:>9.3f}")
    print("=" * 70)

    write_host_summary(results)

    make_scatter(results, "mean_rtt", "Average RTT (ms)",
                 "Section 3.2 — Average RTT vs Geographic Distance", PNG_RTT)
    make_scatter(results, "jitter", "Jitter σ (ms)",
                 "Section 3.2 — Jitter (sample SD) vs Geographic Distance", PNG_JIT)
